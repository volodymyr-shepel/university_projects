# https://www.kaggle.com/datasets/desalegngeb/students-exam-scores?select=Expanded_data_with_more_features.csv

import argparse
import csv
import sys
import os 
import matplotlib.pyplot as plt
import seaborn as sns
import statistics
import openpyxl
from openpyxl.styles import Font, PatternFill
from pprint import pprint


def read_csv(filename):
    data = {}
    with open(filename) as csv_file:
        csvreader = csv.reader(csv_file)
        headers = next(csvreader)
        headers[0] = "Id"
        for header in headers:
            data.setdefault(header,[])
        for line in csvreader:
            for index in range(len(headers)):
                data[headers[index]].append(float(line[index]) if line[index].isnumeric() else line[index])
        
        data["Id"] = list(range(0,len(data["Id"]))) 
        # there is a problem with ids in this dataset (they are no unique), so I will modify this so they will be unique  
    return data,headers
    


def value_counts(column,description = "Secret"):
    counts = {}
    for row in column:
        counts.setdefault(row,0)
        counts[row] += 1
    counts["Description"] = description
    return counts

def replace_missing_values(data,value_to_replace): # function to handle missing or not collected values
    for key,values in data.items():
        for index in range(len(values)):
            if values[index] == None or values[index] == "":
                data[key][index] = value_to_replace
    return data             
def group_by(data,column,headers):
    groups = {}
    
    for index in range(len(data[column])):
        groups.setdefault(data[column][index],[])
        groups[data[column][index]].append([data[key][index] for key in headers])
    
    return groups

def get_column(groups,column,headers): # do after group_by()
    if column in headers:
        column_index = headers.index(column)
        column_data = {}
        for key in groups.keys():
            column_data.setdefault(key,[])
            for el in groups[key]:
                column_data[key].append(el[column_index])
        return column_data
    else:
        return {}

def statistics_calculation(data,func): # helper function for calculate statistics
    stats = {}
    for key,value in data.items():
        stats[key] = round(func(value),4)
    return stats


def calculate_statistics(data,column_to_group,column_to_calculate,func,headers,description = "Secret"):
    grouped = group_by(data,column_to_group,headers)
    chosen_column = get_column(grouped,column_to_calculate,headers)
    result = statistics_calculation(chosen_column,func)
    result["Description"] = description
    return result


def iloc(data,index,headers):
    element = []
    for key in headers:
        element.append(data[key][index])
    return element


def collect_missing_values(data,description):
    missing_vals = {}
    for key,values in data.items():
        missing_vals.setdefault(key,0)
        for value in values:
            if value == None or value =="":
                missing_vals[key] += 1
    missing_vals = {k:v for k, v in missing_vals.items() if v > 0} # filter dictionary so it returns only columns with missing values
    missing_vals["Description"] = description
    return missing_vals


def is_csv_file(filename):
    with open(filename, 'r',encoding="utf-8") as f:
        try:
            dialect = csv.Sniffer().sniff(f.read(2048))
            return True
        except csv.Error:
            return False

def display(list_do_display):
    for element in list_do_display:
        pprint(element, indent=4)
        print()
    
def write_to_xlsx_file(file,list_to_write):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "main"
    
    row_iterator = 1
    
    font1 = Font(name='Arial', size=12, bold=True,color = "C7395F")
    font2 = Font(name='Arial', size=12, bold=True,color = "C7395F")
    # create a fill object with the desired color
    color1  = PatternFill(start_color='DED4E8', end_color='DED4E8', fill_type='solid')
    color2  = PatternFill(start_color='FCA311', end_color='FCA311', fill_type='solid')

    for element in list_to_write:
        for k,v in element.items():
            column_iterator = 1
            cell1 = sheet.cell(column = column_iterator,row = row_iterator,value = k)
            cell1.font = font1
            cell1.fill = color1
            column_iterator+=1
            cell2 = sheet.cell(column = column_iterator,row = row_iterator,value = v)
            cell2.fill = color2
            cell2.font = font2
            row_iterator += 1
        row_iterator += 2
    workbook.save(file)



def main(file_to_read):
    data,_headers = read_csv(file_to_read)
    
    
    missang_vals = collect_missing_values(data,"Missing values in columns") 
    data = replace_missing_values(data,"Not collected") # handle missing values
    
    gender_writing_score_means = calculate_statistics(data,"Gender","WritingScore",statistics.mean,_headers,"Average writing score based on gender") 
    gender_reading_score_medians = calculate_statistics(data,"Gender","ReadingScore",statistics.median,_headers,"Median reading score based on gender") 
    ethnic_group_math_score_means = calculate_statistics(data,"EthnicGroup","MathScore",statistics.mean,_headers,"Average math score based on ethnic group") 
    study_hours_score_means = calculate_statistics(data,"WklyStudyHours","MathScore",statistics.mean,_headers,"Average math score based on studying hours") 
    parents_education_srore_means = calculate_statistics(data,"ParentEduc","MathScore",statistics.mean,_headers,"Average math score based on parents' education") 
    gender_math_score = calculate_statistics(data,"Gender","MathScore",statistics.mean,_headers,"Average math score based on parents' education")
     
    gender_value_counts = value_counts(data["Gender"],"Gender value counts")
    study_hours_value_counts = value_counts(data["WklyStudyHours"],"Study hours value counts")
    
    resulted_list = [missang_vals,gender_writing_score_means,gender_reading_score_medians,ethnic_group_math_score_means,study_hours_score_means,gender_value_counts,study_hours_value_counts,parents_education_srore_means,gender_math_score]    
    
    return resulted_list
    
    
    
if __name__ == "__main__":
    
    parser = argparse.ArgumentParser(description="Read and process csv file")   
    parser.add_argument("--path",type = str,help = "Path of the csv file to process")
    parser.add_argument('-o', type=str,help = "Enables the option to save the output to provided the excel file")
    
    # parser.add_argument('x', type=int, help='The first value to multiply')
    args = parser.parse_args()
    
    
    if not (args.path and os.path.isfile(args.path) and is_csv_file(args.path)):
        sys.exit("Provided file does not exist")
    
    
    
        
    result = main(args.path)
    
    if args.o:
        if os.path.exists(args.o):
            sys.exit("The file with provided xlsx file already exists")
        write_to_xlsx_file(args.o,result)
    else:
        display(result)
        
# python app8.py --path students_test.csv
# python app8.py --path students_test.csv -o analysis.xlsx
        
    
    
